import time
import traceback, sys
from openpyxl import Workbook, load_workbook
import json
import os
import datetime
import sys
from brewserver import settings



class Temp_profile:
    def __init__(self, queue):
        self.queue = queue
        self.filename = None
        self.running = False
        self.previous_step = {"time": datetime.datetime.now(), 'temp': None}
        self.next_step = {"time": datetime.datetime.now(), 'temp': None}
        self.next_duration = None

    def getFile(self):
        return self.filename

    def open(self, filename):

        directory=os.path.join(settings.BASE_DIR, 'temp_profiles')
        path = os.path.join(directory, filename)
        wb = load_workbook(path)

        self.filename = filename
        ws = wb.active
        self.tab = [[x[0].value, x[1].value] for x in zip(ws['A'], ws['B']) if x[0].value is not None]

    def run(self):
        if not self.running:
            self.running = True
            try:
                print("Start Temperature Profile")
                for i, d in enumerate(self.tab):
                    minutes, temp = d
                    print(i, minutes, temp)
                    try:
                        minutes = float(minutes)
                        temp = float(temp)
                    except ValueError:
                        print("ValueError")
                        continue
                    msg = {"rampup2": {"minutes": minutes, "temp": temp}}
                    self.previous_step["time"] = datetime.datetime.now()
                    self.previous_step["temp"] = temp
                    self.next_step["time"] = self.previous_step["time"] + datetime.timedelta(minutes=minutes)
                    try:
                        self.next_step["temp"] = self.tab[i+1][1]
                        self.next_duration = self.tab[i+1][0]
                    except IndexError:
                        self.next_step["temp"] = None
                        self.next_duration = None

                    self.queue.put(json.dumps(msg))
                    if (minutes != 0):
                        time.sleep(minutes * 60)
                        msg = {"rampup2": {"minutes": 0, "temp": temp}}
                        self.queue.put(json.dumps(msg))
            except:
                self.running = False
            self.running = False



    def status(self):
        pass

    def status(self):
        dur=None
        if self.next_duration is not None:
            dur=int(self.next_duration*60)
        return {
            "current_temp": self.previous_step["temp"],
            "next_temp": self.next_step["temp"],
            "current_time": (self.next_step["time"]-datetime.datetime.now()).seconds,
            "next_duration": dur
        }


if __name__ == "__main__":
    tp = Temp_profile(None)
    tp.open("temperaturprofil.xlsx")
    tp.run()

